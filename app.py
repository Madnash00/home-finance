#!/usr/bin/env python3
import base64, csv, hashlib, hmac, io, json, mimetypes, os, re, shutil, sqlite3, sys, tempfile, threading, unicodedata
from datetime import datetime, timedelta
from http.server import ThreadingHTTPServer, SimpleHTTPRequestHandler
from pathlib import Path
from urllib.parse import parse_qs, urlparse

ROOT=Path(__file__).parent
env_file=ROOT/'.env'
if env_file.exists():
 for line in env_file.read_text(encoding='utf-8').splitlines():
  line=line.strip()
  if line and not line.startswith('#') and '=' in line:
   key,value=line.split('=',1);os.environ.setdefault(key.strip(),value.strip())
STATIC=ROOT/'static'; DATA_DIR=Path(os.getenv('DATA_DIR',ROOT/'data'));DATA_DIR.mkdir(parents=True,exist_ok=True);DB=Path(os.getenv('DATABASE_PATH',DATA_DIR/'contabilita.db'))
LEGACY_DB=ROOT/'contabilita.db'
if not DB.exists() and LEGACY_DB.exists() and DB!=LEGACY_DB and not os.getenv('DATA_DIR') and not os.getenv('DATABASE_PATH'):shutil.copy2(LEGACY_DB,DB)
sys.path.insert(0,str(ROOT/'.vendor'))
try:
 import openpyxl
 from pyxlsb import open_workbook
except ImportError:
 openpyxl=None;open_workbook=None
XLSB=Path(os.getenv('INITIAL_MODEL_FILE','')) if os.getenv('INITIAL_MODEL_FILE') else None
XLSX=Path(os.getenv('INITIAL_MOVEMENTS_FILE','')) if os.getenv('INITIAL_MOVEMENTS_FILE') else None

def conn():
 c=sqlite3.connect(DB); c.row_factory=sqlite3.Row; c.execute('PRAGMA foreign_keys=ON'); return c
def migrate():
 c=conn(); c.executescript(SCHEMA)
 cols={x['name'] for x in c.execute('pragma table_info(categories)')}
 if 'year' not in cols:c.execute('alter table categories add column year INTEGER DEFAULT 2026')
 if 'keywords' not in cols:c.execute("alter table categories add column keywords TEXT DEFAULT ''")
 c.execute('update categories set year=2026 where year is null')
 c.execute("update categories set keywords=coalesce((select group_concat(value,'; ') from classification_rules r where r.category_id=categories.id),'') where coalesce(keywords,'')='' and level=3")
 c.commit(); c.close()

def data_maintenance_v3():
 c=conn()
 if c.execute("select value from settings where key='maintenance_v3b'").fetchone():c.close();return
 # Rimuove solo le copie importate che coincidono con una riga storica anche nella descrizione canonica.
 imported=c.execute('select * from transactions where import_batch_id is not null').fetchall(); removed=0
 for n in imported:
  candidates=c.execute("select * from transactions where import_batch_id is null and accounting_date=? and coalesce(value_date,'')=coalesce(?,'') and round(amount,2)=round(?,2) and coalesce(causale,'')=coalesce(?,'')",(n['accounting_date'],n['value_date'],n['amount'],n['causale'])).fetchall()
  new_text=norm(n['combined'])
  if any(norm(x['combined'])==new_text for x in candidates):c.execute('delete from transactions where id=?',(n['id'],));removed+=1
 c.execute("insert into settings(key,value) values('maintenance_v3b',?)",(json.dumps({'removed_import_duplicates':removed}),))
 c.commit();c.close()

def dedupe_classification_rules():
 c=conn();c.execute('delete from classification_rules where id not in (select min(id) from classification_rules group by category_id,lower(trim(value)),operator,field)')
 for cat in c.execute('select id from categories where level=3').fetchall():
  values=[];seen=set()
  for r in c.execute('select value from classification_rules where category_id=? order by priority desc,id',(cat['id'],)).fetchall():
   k=norm(r['value'])
   if k and k not in seen:seen.add(k);values.append(str(r['value']).strip())
  c.execute('update categories set keywords=? where id=?',('; '.join(values),cat['id']))
 c.commit();c.close()

def rebuild_historical_taxonomy():
 c=conn()
 if c.execute("select value from settings where key='taxonomy_2024_2026_v1'").fetchone():c.close();return
 if not XLSB or not XLSB.exists() or open_workbook is None:c.close();return
 source={2024:{},2025:{},2026:{}}
 with open_workbook(XLSB) as wb:
  with wb.get_sheet('DB ANALISI_Consolidato') as sh:
   for i,row in enumerate(sh.rows()):
    if i==0:continue
    v=[x.v for x in row]+[None]*10; causale,detail,kw,year,movement=v[1],v[2],v[5],v[6],v[9]
    try:year=int(year)
    except:continue
    if year not in source or not causale or not detail or movement not in ('ENTRATE','USCITE'):continue
    key=(movement,str(causale),str(detail));source[year].setdefault(key,set())
    if kw and str(kw) not in ('-','None'):source[year][key].add(str(kw).strip())
 for year in (2024,2025,2026):
  # Il 2026 viene aggiornato; 2024/2025 sono ricostruiti una sola volta e poi restano immutabili.
  parents={}
  for (movement,group,name),keywords in sorted(source[year].items()):
   pkey=(movement,group)
   if pkey not in parents:
    p=c.execute('select id from categories where year=? and level=2 and movement=? and name=?',(year,movement,group)).fetchone()
    if not p:
     c.execute('insert into categories(level,code,name,movement,year) values(2,?,?,?,?)',(f'{movement}|{group}|{year}',group,movement,year));pid=c.execute('select last_insert_rowid()').fetchone()[0]
    else:pid=p[0]
    parents[pkey]=pid
   cat=c.execute('select id from categories where year=? and level=3 and movement=? and name=?',(year,movement,name)).fetchone()
   keyword_text='; '.join(sorted(keywords,key=norm))
   if cat:cid=cat[0];c.execute('update categories set parent_id=?,keywords=? where id=?',(parents[pkey],keyword_text,cid))
   else:
    c.execute('insert into categories(parent_id,level,code,name,movement,year,keywords) values(?,3,?,?,?,?,?)',(parents[pkey],f'{movement}|{group}|{name}|{year}',name,movement,year,keyword_text));cid=c.execute('select last_insert_rowid()').fetchone()[0]
   c.execute('delete from classification_rules where category_id=?',(cid,))
   for keyword in sorted(keywords,key=norm):c.execute('insert into classification_rules(name,priority,category_id,field,operator,value,movement,active,stop_processing,confidence) values(?,?,?,?,?,?,?,?,?,?)',(f'Keyword: {keyword}',100,cid,'combined','contains',keyword,movement,1,1,.9))
  # Collega i movimenti storici alle voci dello stesso anno usando il nome della voce precedente.
  for tx in c.execute('select t.id,c.name,c.movement from transactions t join categories c on c.id=t.category_id where t.year=? and c.year!=?',(year,year)).fetchall():
   target=c.execute('select id from categories where year=? and level=3 and movement=? and name=?',(year,tx['movement'],tx['name'])).fetchone()
   if target:c.execute('update transactions set category_id=? where id=?',(target[0],tx['id']))
 c.execute("insert into settings(key,value) values('taxonomy_2024_2026_v1','completed')")
 c.commit();c.close()

def ensure_category_year(year):
 c=conn()
 if year>=2027 and not c.execute('select 1 from categories where year=? limit 1',(year,)).fetchone():
  source=year-1 if c.execute('select 1 from categories where year=? limit 1',(year-1,)).fetchone() else 2026
  mapping={}
  for level in (2,3):
   for old in c.execute('select * from categories where year=? and level=? order by id',(source,level)).fetchall():
    parent=mapping.get(old['parent_id']) if old['parent_id'] else None
    code=f"{old['code']}|{year}"
    c.execute('insert into categories(parent_id,level,code,name,description,movement,sort_order,active,color,include_reports,year,keywords) values(?,?,?,?,?,?,?,?,?,?,?,?)',(parent,level,code,old['name'],old['description'],old['movement'],old['sort_order'],old['active'],old['color'],old['include_reports'],year,old['keywords']))
    mapping[old['id']]=c.execute('select last_insert_rowid()').fetchone()[0]
    if level==3:
     for rule in c.execute('select * from classification_rules where category_id=?',(old['id'],)).fetchall():
      c.execute('insert into classification_rules(name,priority,category_id,field,operator,value,movement,active,stop_processing,confidence) values(?,?,?,?,?,?,?,?,?,?)',(rule['name'],rule['priority'],mapping[old['id']],rule['field'],rule['operator'],rule['value'],rule['movement'],rule['active'],rule['stop_processing'],rule['confidence']))
  c.commit()
 c.close()
def norm(s): return re.sub(r'\s+',' ',unicodedata.normalize('NFKD',str(s or '')).encode('ascii','ignore').decode().upper()).strip()
def xl_date(v):
 if v is None or v=='': return None
 if isinstance(v,datetime): return v.date().isoformat()
 if isinstance(v,(float,int)): return (datetime(1899,12,30)+timedelta(days=v)).date().isoformat()
 s=str(v).replace('T04:00:00','');
 for f in ('%Y-%m-%d','%d.%m.%Y','%d/%m/%Y'):
  try:return datetime.strptime(s,f).date().isoformat()
  except:pass
 return s[:10]
def fp(account,date,value,desc,ext='',channel=''): return hashlib.sha256('|'.join(map(norm,[account,date,f'{value:.2f}',desc,ext,channel])).encode()).hexdigest()

SCHEMA='''
CREATE TABLE IF NOT EXISTS accounts(id INTEGER PRIMARY KEY,banca TEXT,nome TEXT,numero_mascherato TEXT,intestatario TEXT,valuta TEXT DEFAULT 'EUR',saldo_iniziale REAL,saldo_disponibile REAL,fido REAL DEFAULT 0,attivo INTEGER DEFAULT 1,created_at TEXT DEFAULT CURRENT_TIMESTAMP,updated_at TEXT DEFAULT CURRENT_TIMESTAMP);
CREATE TABLE IF NOT EXISTS import_batches(id INTEGER PRIMARY KEY,nome_file TEXT,hash_file TEXT UNIQUE,imported_at TEXT DEFAULT CURRENT_TIMESTAMP,periodo_da TEXT,periodo_a TEXT,account_id INTEGER,righe_lette INTEGER,righe_importate INTEGER,duplicati INTEGER DEFAULT 0,scartate INTEGER DEFAULT 0,anomalie INTEGER DEFAULT 0,stato TEXT,log TEXT);
CREATE TABLE IF NOT EXISTS raw_transactions(id INTEGER PRIMARY KEY,import_batch_id INTEGER,row_number INTEGER,payload TEXT,row_hash TEXT,created_at TEXT DEFAULT CURRENT_TIMESTAMP);
CREATE TRIGGER IF NOT EXISTS raw_no_update BEFORE UPDATE ON raw_transactions BEGIN SELECT RAISE(ABORT,'raw immutabile'); END;
CREATE TRIGGER IF NOT EXISTS raw_no_delete BEFORE DELETE ON raw_transactions BEGIN SELECT RAISE(ABORT,'raw immutabile'); END;
CREATE TABLE IF NOT EXISTS categories(id INTEGER PRIMARY KEY,parent_id INTEGER,level INTEGER,code TEXT UNIQUE,name TEXT,description TEXT,movement TEXT,sort_order INTEGER DEFAULT 0,active INTEGER DEFAULT 1,color TEXT DEFAULT '#2563eb',include_reports INTEGER DEFAULT 1);
CREATE TABLE IF NOT EXISTS classification_rules(id INTEGER PRIMARY KEY,name TEXT,priority INTEGER,category_id INTEGER,field TEXT,operator TEXT,value TEXT,movement TEXT,active INTEGER DEFAULT 1,stop_processing INTEGER DEFAULT 1,confidence REAL DEFAULT .85,uses INTEGER DEFAULT 0,last_used TEXT);
CREATE TABLE IF NOT EXISTS transactions(id INTEGER PRIMARY KEY,account_id INTEGER,import_batch_id INTEGER,accounting_date TEXT,value_date TEXT,causale TEXT,description TEXT,extended_description TEXT,combined TEXT,credit REAL,debit REAL,amount REAL,currency TEXT DEFAULT 'EUR',balance REAL,channel TEXT,month INTEGER,year INTEGER,movement TEXT,category_id INTEGER,classification_rule_id INTEGER,classification_mode TEXT DEFAULT 'automatica',confidence REAL DEFAULT 0,verified INTEGER DEFAULT 0,note TEXT,excluded INTEGER DEFAULT 0,fingerprint TEXT UNIQUE,possible_duplicate INTEGER DEFAULT 0,created_at TEXT DEFAULT CURRENT_TIMESTAMP,updated_at TEXT DEFAULT CURRENT_TIMESTAMP);
CREATE TABLE IF NOT EXISTS budget_entries(id INTEGER PRIMARY KEY,year INTEGER,month INTEGER,category_id INTEGER,amount REAL,kind TEXT,scenario TEXT DEFAULT 'Base',version INTEGER DEFAULT 1,status TEXT DEFAULT 'pubblicato',method TEXT,notes TEXT,updated_at TEXT DEFAULT CURRENT_TIMESTAMP);
CREATE TABLE IF NOT EXISTS loans(id INTEGER PRIMARY KEY,description TEXT,start_date TEXT,installments INTEGER,payment REAL,due_date TEXT,status TEXT,notes TEXT,paid INTEGER DEFAULT 0,lender TEXT);
CREATE TABLE IF NOT EXISTS audit_log(id INTEGER PRIMARY KEY,actor TEXT DEFAULT 'local',created_at TEXT DEFAULT CURRENT_TIMESTAMP,entity TEXT,entity_id INTEGER,operation TEXT,before_value TEXT,after_value TEXT,reason TEXT,origin TEXT);
CREATE TABLE IF NOT EXISTS settings(key TEXT PRIMARY KEY,value TEXT);
CREATE INDEX IF NOT EXISTS ix_tx_date ON transactions(accounting_date); CREATE INDEX IF NOT EXISTS ix_tx_cat ON transactions(category_id); CREATE INDEX IF NOT EXISTS ix_tx_year_month ON transactions(year,month);
'''

def seed():
 if DB.exists() and conn().execute('select count(*) from transactions').fetchone()[0]>0:return
 c=conn(); c.executescript(SCHEMA)
 if not XLSB or not XLSB.exists() or open_workbook is None:c.commit();c.close();return
 catmap={}; rules=[]; budgets=[]; loans=[]
 with open_workbook(XLSB) as wb:
  with wb.get_sheet('DB ANALISI_Consolidato') as sh:
   for i,row in enumerate(sh.rows()):
    v=[x.v for x in row]
    if i==0 or len(v)<10:continue
    group,causale,detail,desc,keyword,kwlist,year,month,_,movement=(v+[None]*10)[:10]
    if not causale or not detail or movement not in ('ENTRATE','USCITE'):continue
    parent_key=(movement,str(causale));
    if parent_key not in catmap:
     code='|'.join(parent_key); c.execute('insert or ignore into categories(level,code,name,movement) values(2,?,?,?)',(code,causale,movement)); catmap[parent_key]=c.execute('select id from categories where code=?',(code,)).fetchone()[0]
    key=(movement,str(causale),str(detail))
    if key not in catmap:
     code='|'.join(key); c.execute('insert or ignore into categories(parent_id,level,code,name,description,movement) values(?,3,?,?,?,?)',(catmap[parent_key],code,detail,desc,movement)); catmap[key]=c.execute('select id from categories where code=?',(code,)).fetchone()[0]
    if kwlist and str(kwlist) not in ('-','None'):
     rules.append((f'{detail}: {kwlist}',100,catmap[key],'combined','contains',str(kwlist),movement,1,.9))
  with wb.get_sheet('DB ANALISI_Forecast') as sh:
   for i,row in enumerate(sh.rows()):
    v=[x.v for x in row]
    if i==0 or len(v)<8 or v[0] not in ('ENTRATE','USCITE'):continue
    key=(v[0],str(v[2]),str(v[3])); cid=catmap.get(key)
    if cid:
     budgets += [(2026,None,cid,v[4] or 0,'budget','Base',1,'pubblicato','Excel',None),(2026,None,cid,v[6] or 0,'forecast','Base',1,'pubblicato','Excel',None)]
  with wb.get_sheet('FINANZIAMENTI') as sh:
   for i,row in enumerate(sh.rows()):
    v=[x.v for x in row]
    if i and v and v[0]: loans.append((v[0],xl_date(v[1]),int(v[2] or 0),v[3] or 0,xl_date(v[4]),v[5],v[6],0,v[6]))
 c.executemany('insert into classification_rules(name,priority,category_id,field,operator,value,movement,stop_processing,confidence) values(?,?,?,?,?,?,?,?,?)',rules)
 c.executemany('insert into budget_entries(year,month,category_id,amount,kind,scenario,version,status,method,notes) values(?,?,?,?,?,?,?,?,?,?)',budgets)
 c.executemany('insert into loans(description,start_date,installments,payment,due_date,status,notes,paid,lender) values(?,?,?,?,?,?,?,?,?)',loans)
 # historical transactions
 with open_workbook(XLSB) as wb:
  with wb.get_sheet('DB MOVIMENTI') as sh:
   for i,row in enumerate(sh.rows()):
    if i==0:continue
    v=[x.v for x in row]+[None]*14
    if not v[0] or v[6] is None:continue
    bank=str(v[0]); aid=c.execute('select id from accounts where banca=?',(bank,)).fetchone()
    if not aid: c.execute('insert into accounts(banca,nome,numero_mascherato,saldo_iniziale) values(?,?,?,?)',(bank,f'Conto {bank}','••••',v[8] or 0)); aid=c.execute('select last_insert_rowid()').fetchone()[0]
    else: aid=aid[0]
    date=xl_date(v[1]); amount=float(v[6]); text=str(v[5] or v[4] or '')
    match=c.execute("select id,category_id,confidence from classification_rules where active=1 and ? like '%'||upper(value)||'%' order by priority desc limit 1",(norm(text),)).fetchone()
    f=fp(bank,date,amount,text)
    c.execute('insert or ignore into transactions(account_id,accounting_date,value_date,causale,description,combined,amount,balance,month,year,movement,category_id,classification_rule_id,confidence,fingerprint) values(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)',(aid,date,xl_date(v[2]),v[3],v[4],text,amount,v[8],int(v[9] or 0),int(v[10] or 0),v[11],match['category_id'] if match else None,match['id'] if match else None,match['confidence'] if match else 0,f))
 c.commit(); c.close()
 if XLSX and XLSX.exists():import_xlsx(XLSX,True)

def drive_backup(reason='manuale'):
 target=os.getenv('DRIVE_BACKUP_DIR','').strip()
 if not target:return {'configured':False,'message':'DRIVE_BACKUP_DIR non configurata'}
 folder=Path(target).expanduser();folder.mkdir(parents=True,exist_ok=True);stamp=datetime.now().strftime('%Y%m%d-%H%M%S');name=f'contabilita-{stamp}.db';dest=folder/name
 source=conn();backup=sqlite3.connect(dest)
 try:source.backup(backup)
 finally:backup.close();source.close()
 latest=folder/'contabilita-latest.db';shutil.copy2(dest,latest)
 meta={'configured':True,'file':str(dest),'latest':str(latest),'reason':reason,'created_at':datetime.now().isoformat(timespec='seconds')};(folder/'contabilita-latest.json').write_text(json.dumps(meta,ensure_ascii=False,indent=2),encoding='utf-8');return meta

def auto_backup(reason):
 if os.getenv('DRIVE_AUTO_BACKUP','true').lower() in ('1','true','yes') and os.getenv('DRIVE_BACKUP_DIR'):
  try:return drive_backup(reason)
  except Exception as e:return {'configured':True,'error':str(e)}
 return None

def import_xlsx(path,seed_mode=False):
 if openpyxl is None:raise RuntimeError('Dipendenze Excel mancanti: eseguire pip install -r requirements.txt')
 data=Path(path).read_bytes(); h=hashlib.sha256(data).hexdigest(); c=conn()
 old=c.execute('select * from import_batches where hash_file=?',(h,)).fetchone()
 if old:return dict(old)
 wb=openpyxl.load_workbook(io.BytesIO(data),data_only=True); sh=wb.active; header=None
 for i,row in enumerate(sh.iter_rows(values_only=True),1):
  vals=[norm(x) for x in row]
  if 'DATA CONTABILE' in vals and 'DESCRIZIONE' in vals: header=i; heads=vals; break
 if not header: raise ValueError('Intestazione movimenti non rilevata')
 meta={}
 for row in sh.iter_rows(min_row=1,max_row=header-1,values_only=True):
  vals=list(row)
  for j,x in enumerate(vals):
   k=norm(x)
   if k.endswith(':') and j+1<len(vals):meta[k]=vals[j+1:j+3]
 num=str((meta.get('NUMERO CONTO:') or [''])[0]); masked='•••• '+num[-4:] if num else '••••'; owner=str((meta.get('INTESTATARIO CONTO:') or [''])[0])
 init=(meta.get('SALDO CONTABILE INIZIALE AL:') or [None,0]); final=(meta.get('SALDO CONTABILE FINALE AL:') or [None,0]); avail=(meta.get('SALDO DISPONIBILE AL (ESCLUSO FIDO):') or [None,0])
 bank='ISP'; acc=c.execute('select id from accounts where banca=?',(bank,)).fetchone()
 if not acc:c.execute('insert into accounts(banca,nome,numero_mascherato,intestatario,saldo_iniziale,saldo_disponibile) values(?,?,?,?,?,?)',(bank,'Conto principale ISP',masked,owner,float(init[1] or 0),float(avail[1] or 0))); aid=c.execute('select last_insert_rowid()').fetchone()[0]
 else:aid=acc[0]; c.execute('update accounts set numero_mascherato=?,intestatario=?,saldo_iniziale=?,saldo_disponibile=? where id=?',(masked,owner,float(init[1] or 0),float(avail[1] or 0),aid))
 c.execute('insert into import_batches(nome_file,hash_file,periodo_da,periodo_a,account_id,righe_lette,righe_importate,stato,log) values(?,?,?,?,?,0,0,?,?)',(Path(path).name,h,xl_date(init[0]),xl_date(final[0]),aid,'in corso','{}')); bid=c.execute('select last_insert_rowid()').fetchone()[0]
 imported=dups=invalid=uncat=0
 for rn,row in enumerate(sh.iter_rows(min_row=header+1,values_only=True),header+1):
  v=list(row)+[None]*8
  if norm(v[2]).startswith('SALDO CONTABILE INIZIALE'):continue
  if not v[0] or (v[3] is None and v[4] is None): invalid+=1; continue
  amount=float(v[3] or v[4] or 0); amount=abs(amount) if v[3] else -abs(amount); date=xl_date(v[0]); text=' '.join(filter(None,[str(v[2] or ''),str(v[5] or '')])); channel=str(v[6] or '')
  raw=json.dumps([str(x) if x is not None else None for x in row],ensure_ascii=False); rh=hashlib.sha256(raw.encode()).hexdigest(); c.execute('insert into raw_transactions(import_batch_id,row_number,payload,row_hash) values(?,?,?,?)',(bid,rn,raw,rh))
  f=fp(bank,date,amount,v[2],v[5],channel)
  canonical=norm(text); candidates=c.execute("select combined,causale from transactions where account_id=? and accounting_date=? and coalesce(value_date,'')=coalesce(?,'') and round(amount,2)=round(?,2) and coalesce(causale,'')=coalesce(?,'')",(aid,date,xl_date(v[1]),amount,v[2])).fetchall()
  if c.execute('select 1 from transactions where fingerprint=?',(f,)).fetchone() or any(norm(x['combined'])==canonical for x in candidates):dups+=1;continue
  nt=norm(text); match=c.execute("select id,category_id,confidence from classification_rules where active=1 and movement=? and ? like '%'||upper(value)||'%' order by priority desc,length(value) desc limit 1",('ENTRATE' if amount>0 else 'USCITE',nt)).fetchone()
  if not match:uncat+=1
  c.execute('insert into transactions(account_id,import_batch_id,accounting_date,value_date,causale,description,extended_description,combined,credit,debit,amount,channel,month,year,movement,category_id,classification_rule_id,confidence,fingerprint) values(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)',(aid,bid,date,xl_date(v[1]),v[2],v[2],v[5],text,abs(amount) if amount>0 else 0,abs(amount) if amount<0 else 0,amount,channel,int(date[5:7]),int(date[:4]),'ENTRATE' if amount>0 else 'USCITE',match['category_id'] if match else None,match['id'] if match else None,match['confidence'] if match else 0,f)); imported+=1
 c.execute('update import_batches set righe_lette=?,righe_importate=?,duplicati=?,scartate=?,anomalie=?,stato=?,log=? where id=?',(imported+dups+invalid,imported,dups,invalid,uncat,'completato',json.dumps({'non_categorizzati':uncat,'saldo_finale_dichiarato':final[1]}),bid)); c.commit(); out=dict(c.execute('select * from import_batches where id=?',(bid,)).fetchone()); c.close(); auto_backup('importazione movimenti'); return out

def rows(sql,args=()):
 c=conn(); out=[dict(x) for x in c.execute(sql,args).fetchall()]; c.close(); return out
def overview(year=2026):
 c=conn(); k=dict(c.execute('select coalesce(sum(amount),0) net,coalesce(sum(case when amount>0 then amount end),0) income,coalesce(sum(case when amount<0 then amount end),0) expense,count(*) count,sum(category_id is null) uncat from transactions where year=? and excluded=0',(year,)).fetchone())
 balances=[dict(x) for x in c.execute('select banca,nome,saldo_disponibile from accounts where attivo=1')]
 monthly=[dict(x) for x in c.execute('select month,round(sum(case when amount>0 then amount else 0 end),2) income,round(sum(case when amount<0 then amount else 0 end),2) expense,round(sum(amount),2) net from transactions where year=? and excluded=0 group by month order by month',(year,))]
 top=[dict(x) for x in c.execute('select coalesce(c.name,"USCITE VARIE") category,round(sum(t.amount),2) amount,count(*) count from transactions t left join categories c on c.id=t.category_id where t.year=? and t.amount<0 and t.excluded=0 group by coalesce(c.name,"USCITE VARIE") order by sum(t.amount) asc limit 8',(year,))]
 top_income=[dict(x) for x in c.execute('select coalesce(c.name,"ENTRATE VARIE") category,round(sum(t.amount),2) amount,count(*) count from transactions t left join categories c on c.id=t.category_id where t.year=? and t.amount>0 and t.excluded=0 group by coalesce(c.name,"ENTRATE VARIE") order by sum(t.amount) desc limit 8',(year,))]
 bf=[dict(x) for x in c.execute("select kind,round(sum(amount),2) amount from budget_entries where year=? group by kind",(year,))]
 c.close(); return {'kpi':k,'balances':balances,'monthly':monthly,'top_expenses':top,'top_income':top_income,'planning':bf}

def planning(year=2026):
 ensure_category_year(year); c=conn(); month=datetime.now().month if year==datetime.now().year else 12
 sql='''WITH actuals AS (
   SELECT category_id,
    SUM(CASE WHEN year=? THEN amount ELSE 0 END) actual,
    SUM(CASE WHEN year=? AND month<=? THEN amount ELSE 0 END) actual_ytd,
    SUM(CASE WHEN year=? AND month<=? THEN amount ELSE 0 END) ly,
    SUM(CASE WHEN year=? AND month<=? THEN amount ELSE 0 END) lly
   FROM transactions WHERE excluded=0 AND year BETWEEN ? AND ? GROUP BY category_id
  ), plans AS (
   SELECT category_id,
    SUM(CASE WHEN kind='budget' THEN amount ELSE 0 END) budget,
    SUM(CASE WHEN kind='forecast' THEN amount ELSE 0 END) forecast
   FROM budget_entries WHERE year=? AND status!='archiviato' GROUP BY category_id
  )
  SELECT c.id category_id,c.movement,coalesce(p.name,'—') gruppo,c.name categoria,
   round(coalesce(a.actual,0),2) actual,round(coalesce(a.actual_ytd,0),2) actual_ytd,
   round(coalesce(a.ly,0),2) ly,round(coalesce(a.lly,0),2) lly,
   round(coalesce(pl.budget,0),2) budget,round(coalesce(pl.forecast,0),2) forecast
  FROM categories c LEFT JOIN categories p ON p.id=c.parent_id
  LEFT JOIN actuals a ON a.category_id=c.id LEFT JOIN plans pl ON pl.category_id=c.id
  WHERE c.level=3 AND c.active=1 AND c.year=? AND (a.category_id IS NOT NULL OR pl.category_id IS NOT NULL)
  ORDER BY c.movement,p.name,c.sort_order,c.name'''
 args=(year,year,month,year-1,month,year-2,month,year-2,year,year,year)
 out=[]
 for x in c.execute(sql,args):
  r=dict(x); r['delta_ly']=round(r['actual_ytd']-r['ly'],2); r['delta_lly']=round(r['actual_ytd']-r['lly'],2); r['delta_budget']=round(r['actual_ytd']-r['budget'],2); r['delta_forecast']=round(r['actual_ytd']-r['forecast'],2)
  r['pct_ly']=round((r['delta_ly']/abs(r['ly'])*100),1) if r['ly'] else None
  r['pct_lly']=round((r['delta_lly']/abs(r['lly'])*100),1) if r['lly'] else None
  r['pct_budget']=round((r['delta_budget']/abs(r['budget'])*100),1) if r['budget'] else None
  r['pct_forecast']=round((r['delta_forecast']/abs(r['forecast'])*100),1) if r['forecast'] else None
  out.append(r)
 c.close(); return out

def save_entity(entity,body):
 tables={'planning':('budget_entries',('year','month','category_id','amount','kind','status','notes')),'loans':('loans',('description','start_date','payment','due_date','status','notes','lender')),'categories':('categories',('parent_id','level','code','name','description','movement','sort_order','active','color','include_reports','year','keywords'))}
 if entity not in tables: raise ValueError('Entità non modificabile')
 table,fields=tables[entity]; c=conn(); eid=body.pop('id',None); clean={k:body[k] for k in fields if k in body}
 if not clean: raise ValueError('Nessun campo valido')
 before=None
 if eid:
  old=c.execute(f'select * from {table} where id=?',(eid,)).fetchone(); before=dict(old) if old else None
  if not old: raise ValueError('Elemento non trovato')
  if table=='budget_entries': clean['version']=int(old['version'] or 1)+1
  c.execute(f"update {table} set "+','.join(f'{k}=?' for k in clean)+(" ,updated_at=CURRENT_TIMESTAMP" if table=='budget_entries' else '')+' where id=?',(*clean.values(),eid))
 else:
  cols=','.join(clean); c.execute(f'insert into {table}({cols}) values({",".join("?" for _ in clean)})',tuple(clean.values())); eid=c.execute('select last_insert_rowid()').fetchone()[0]
 if table=='categories' and 'keywords' in clean:
  c.execute('delete from classification_rules where category_id=?',(eid,))
  movement=clean.get('movement') or c.execute('select movement from categories where id=?',(eid,)).fetchone()[0]
  unique_keywords=list({norm(x):x.strip() for x in re.split(r'[;\n]',clean['keywords'] or '') if x.strip()}.values());clean['keywords']='; '.join(unique_keywords);c.execute('update categories set keywords=? where id=?',(clean['keywords'],eid))
  for keyword in unique_keywords:
   c.execute('insert into classification_rules(name,priority,category_id,field,operator,value,movement,active,stop_processing,confidence) values(?,?,?,?,?,?,?,?,?,?)',(f'Keyword: {keyword}',100,eid,'combined','contains',keyword,movement,1,1,.9))
 c.execute('insert into audit_log(entity,entity_id,operation,before_value,after_value,reason,origin) values(?,?,?,?,?,?,?)',(table,eid,'UPDATE' if before else 'CREATE',json.dumps(before,default=str),json.dumps(clean,default=str),'Modifica utente','web'))
 c.commit(); out=dict(c.execute(f'select * from {table} where id=?',(eid,)).fetchone()); c.close(); auto_backup(f'modifica {table}'); return out

class Handler(SimpleHTTPRequestHandler):
 def authenticated(self):
  password=os.getenv('APP_PASSWORD','')
  if not password:return True
  header=self.headers.get('Authorization','')
  if not header.startswith('Basic '):return False
  try:user,supplied=base64.b64decode(header[6:]).decode().split(':',1)
  except:return False
  return hmac.compare_digest(user,os.getenv('APP_USERNAME','admin')) and hmac.compare_digest(supplied,password)
 def require_auth(self):
  if self.authenticated():return False
  self.send_response(401);self.send_header('WWW-Authenticate','Basic realm="Casa Finance"');self.send_header('Cache-Control','no-store');self.end_headers();return True
 def send_json(self,obj,status=200):
  b=json.dumps(obj,ensure_ascii=False,default=str).encode(); self.send_response(status); self.send_header('Content-Type','application/json; charset=utf-8');self.send_header('Content-Length',len(b));self.send_header('X-Content-Type-Options','nosniff');self.send_header('X-Frame-Options','DENY');self.send_header('Referrer-Policy','no-referrer');self.send_header('Cache-Control','no-store');self.end_headers();self.wfile.write(b)
 def do_GET(self):
  p=urlparse(self.path); q=parse_qs(p.query)
  try:
   if p.path=='/health':return self.send_json({'status':'ok'})
   if self.require_auth():return
   if p.path=='/api/overview':return self.send_json(overview(int(q.get('year',[2026])[0])))
   if p.path=='/api/transactions':
    y=int(q.get('year',[2026])[0]); search=q.get('search',[''])[0]; limit=min(int(q.get('limit',[100])[0]),500); offset=int(q.get('offset',[0])[0]); like=f'%{search}%'
    return self.send_json(rows('select t.id,a.banca,t.accounting_date,t.value_date,t.causale,t.extended_description,t.amount,t.balance,t.movement,t.category_id,coalesce(c.name,"Da classificare") category,t.confidence,t.verified,t.note from transactions t join accounts a on a.id=t.account_id left join categories c on c.id=t.category_id where t.year=? and (t.combined like ? or c.name like ?) order by t.accounting_date desc,t.id desc limit ? offset ?',(y,like,like,limit,offset)))
   if p.path=='/api/transactions/count':
    y=int(q.get('year',[2026])[0]);search=q.get('search',[''])[0];like=f'%{search}%';c=conn();n=c.execute('select count(*) from transactions t left join categories c on c.id=t.category_id where t.year=? and (t.combined like ? or c.name like ?)',(y,like,like)).fetchone()[0];c.close();return self.send_json({'total':n})
   if p.path=='/api/analysis':
    y=int(q.get('year',[2026])[0]);return self.send_json(planning(y))
   if p.path=='/api/planning':
    y=int(q.get('year',[2026])[0]);return self.send_json(rows('select b.*,c.name category,c.movement,p.name gruppo from budget_entries b join categories c on c.id=b.category_id left join categories p on p.id=c.parent_id where b.year=? order by c.movement,p.name,c.name,b.kind',(y,)))
   if p.path=='/api/analysis/transactions':
    y=int(q.get('year',[2026])[0]); cid=int(q.get('category_id',[0])[0]); return self.send_json(rows('select t.id,t.accounting_date,t.value_date,t.extended_description,t.amount,t.balance,a.banca from transactions t join accounts a on a.id=t.account_id where t.year=? and t.category_id=? and t.excluded=0 order by t.accounting_date desc',(y,cid)))
   if p.path=='/api/monthly':return self.send_json(overview(int(q.get('year',[2026])[0]))['monthly'])
   if p.path=='/api/categories':
    y=int(q.get('year',[2026])[0]);ensure_category_year(y);return self.send_json(rows('select c.*,p.name parent_name,(select count(*) from transactions where category_id=c.id) uses from categories c left join categories p on p.id=c.parent_id where c.year=? order by c.movement,c.sort_order,c.name',(y,)))
   if p.path=='/api/rules':return self.send_json(rows('select r.*,c.name category from classification_rules r join categories c on c.id=r.category_id order by priority desc,id'))
   if p.path=='/api/loans':
    status=q.get('status',['TUTTI'])[0].upper(); c=conn(); last=c.execute('select max(accounting_date) from transactions').fetchone()[0] or datetime.now().date().isoformat(); data=[]
    for x in c.execute('select * from loans order by status desc,due_date'):
     r=dict(x); start=datetime.fromisoformat(r['start_date']).date() if r['start_date'] else None; due=datetime.fromisoformat(r['due_date']).date() if r['due_date'] else None; end=datetime.fromisoformat(last).date()
     r['installments']=max(0,(due.year-start.year)*12+due.month-start.month+1) if start and due else 0
     r['paid']=min(r['installments'],max(0,(end.year-start.year)*12+end.month-start.month+1)) if start else 0
     if status=='TUTTI' or norm(r['status'])==status:data.append(r)
    c.close(); return self.send_json(data)
   if p.path=='/api/imports':return self.send_json(rows('select b.*,a.banca from import_batches b left join accounts a on a.id=b.account_id order by imported_at desc'))
   if p.path=='/api/audit':return self.send_json(rows('select * from audit_log order by created_at desc limit 200'))
   if p.path=='/api/alerts':
    o=overview(2026); alerts=[]
    if o['kpi']['uncat']:alerts.append({'severity':'warning','title':'Movimenti da classificare','description':f"{o['kpi']['uncat']} movimenti richiedono revisione",'action':'Apri Movimenti'})
    for x in o['top_expenses'][:3]:alerts.append({'severity':'info','title':f"Focus {x['category']}",'description':f"{x['count']} movimenti, totale € {abs(x['amount']):,.2f}",'action':'Analizza dettaglio'})
    return self.send_json(alerts)
   if p.path=='/api/export/csv':
    data=rows('select t.accounting_date,a.banca,t.causale,t.extended_description,t.amount,t.movement,c.name categoria from transactions t join accounts a on a.id=t.account_id left join categories c on c.id=t.category_id order by t.accounting_date'); s=io.StringIO();w=csv.DictWriter(s,fieldnames=data[0].keys());w.writeheader();w.writerows(data);b=s.getvalue().encode('utf-8-sig');self.send_response(200);self.send_header('Content-Type','text/csv');self.send_header('Content-Disposition','attachment; filename=movimenti.csv');self.end_headers();return self.wfile.write(b)
   if p.path=='/api/backup':
    b=DB.read_bytes();self.send_response(200);self.send_header('Content-Type','application/octet-stream');self.send_header('Content-Disposition','attachment; filename=contabilita-backup.db');self.end_headers();return self.wfile.write(b)
   if p.path=='/api/drive/status':return self.send_json({'configured':bool(os.getenv('DRIVE_BACKUP_DIR')),'directory':os.getenv('DRIVE_BACKUP_DIR',''),'automatic':os.getenv('DRIVE_AUTO_BACKUP','true').lower() in ('1','true','yes')})
   if p.path=='/':self.path='/index.html'
   return super().do_GET()
  except Exception as e:return self.send_json({'error':str(e)},500)
 def do_POST(self):
  try:
   if self.require_auth():return
   if self.path=='/api/import':
    n=int(self.headers.get('Content-Length',0)); data=self.rfile.read(n); tmp=ROOT/'uploaded.xlsx';tmp.write_bytes(data); out=import_xlsx(tmp);tmp.unlink();return self.send_json(out,201)
   if self.path in ('/api/planning','/api/loans','/api/categories'):
    body=json.loads(self.rfile.read(int(self.headers.get('Content-Length',0))) or b'{}'); return self.send_json(save_entity(self.path.split('/')[-1],body),201)
   if self.path.startswith('/api/transactions/'):
    tid=int(self.path.split('/')[-1]); body=json.loads(self.rfile.read(int(self.headers.get('Content-Length',0))) or b'{}'); c=conn();before=dict(c.execute('select * from transactions where id=?',(tid,)).fetchone()); allowed={('note' if k=='notes' else k):v for k,v in body.items() if k in ('category_id','verified','note','notes','excluded')};
    if allowed:
     c.execute('update transactions set '+','.join(f'{k}=?' for k in allowed)+',classification_mode="manuale",updated_at=CURRENT_TIMESTAMP where id=?',(*allowed.values(),tid));c.execute('insert into audit_log(entity,entity_id,operation,before_value,after_value,origin) values(?,?,?,?,?,?)',('transactions',tid,'UPDATE',json.dumps(before,default=str),json.dumps(allowed),'web'))
    c.commit();c.close();auto_backup('riclassificazione movimento');return self.send_json({'ok':True})
   if self.path=='/api/drive/backup':return self.send_json(drive_backup('backup richiesto dall app'),201)
   return self.send_json({'error':'not found'},404)
  except Exception as e:return self.send_json({'error':str(e)},400)
 def translate_path(self,path): return str(STATIC/Path(urlparse(path).path).name)
 def log_message(self,fmt,*args): pass

if __name__=='__main__':
 migrate(); seed(); rebuild_historical_taxonomy(); data_maintenance_v3(); dedupe_classification_rules(); host=os.getenv('HOST','127.0.0.1');port=int(os.getenv('PORT','8766')); print(f'Contabilità disponibile su http://{host}:{port}'); ThreadingHTTPServer((host,port),Handler).serve_forever()
